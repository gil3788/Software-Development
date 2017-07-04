from openpyxl import Workbook
import datetime

#Global Variables/Objects#
##Required to make reference to the same Excel Spreadsheet##
wb = Workbook()

def reportOverview() :
    wb.create_sheet(index=0, title="READ - Report Overview")
    ws = wb.get_sheet_by_name("READ - Report Overview")
    #Print Titles#
    ws['A1'] = "Doubleclick Billing Report"
    ws['A3'] = "Sheet 1 - DCM Summary Invoices"
    ws['A4'] = "Sheet 2 - DBM Summary Invoices"
    ws['A5'] = "Sheet 3 - DoubleClick Fiscal Invoices"
    ws['A6'] = "Sheet 4 - DBM Matched Invoices "
    ws['A7'] = "Sheet 5 - DCM Matched Invoices "
    ws['B1'] = "Date Generated"
    ws['B2'] = datetime.datetime.now()

def dcmExcelSummary(invoiceObject):
    wb.create_sheet(index=1, title="DCM Summary Invoices")
    ws = wb.get_sheet_by_name("DCM Summary Invoices")
    #Print Titles#
    ws['A1'] = "Invoice Number"
    ws['B1'] = "Date"
    ws['C1'] = "Product"
    ws['D1'] = "Advertiser Name"
    ws['E1'] = "Advertiser ID"
    ws['F1'] = "Transaction ID"
    ws['G1'] = "Campaign Name"
    ws['H1'] = "Campaign ID"
    ws['I1'] = "Campaign Impressions"
    ws['J1'] = "Campaign Clics"
    ws['K1'] = "Invoice Total Amount"
    ws['L1'] = "Invoice Tax Amount"
    ws['M1'] = "Invoice File Name"
    #ws['A2'] = datetime.datetime.now()
    counter = 0
    newIndex = 0
    while counter < len(invoiceObject.keys()) :
        campaignCounter = 0
        if newIndex == 0:
            index = counter + 2 + newIndex
        else:
            index =  2 + newIndex
        ws['A'+str(index)] = counter + 1
        invoiceKey = 'invoice' + str(counter)
        try:
            objectSize = len(invoiceObject['invoice' + str(counter)]["campaignData"]['invoice' + str(counter)].keys())
        except KeyError:
            objectSize = 0


        while campaignCounter < objectSize:
            #print "Campaign Counter: ", campaignCounter
            newIndex = index + campaignCounter
            ws['B'+str(newIndex)] = invoiceObject['invoice'+str(counter)]['date']
            ws['C'+str(newIndex)] =  invoiceObject['invoice'+str(counter)]['product']
            ws['D'+str(newIndex)] =  invoiceObject['invoice'+str(counter)]['advertiserName']
            ws['E'+str(newIndex)] =  invoiceObject['invoice'+str(counter)]['advertiserId']
            ws['F'+str(newIndex)] =  invoiceObject['invoice'+str(counter)]['transactionId']
            ws['G'+str(newIndex)] =  invoiceObject['invoice'+str(counter)]['campaignData']['invoice' + str(counter)]["campaign" + str(campaignCounter)]["campaignName"]
            ws['H'+str(newIndex)] =  invoiceObject['invoice'+str(counter)]['campaignData']['invoice' + str(counter)]["campaign" + str(campaignCounter)]["campaignId"]
            if  invoiceObject['invoice'+str(counter)]['campaignData']['invoice' + str(counter)]["campaign" + str(campaignCounter)]["campaignImp"].find("CPC") ==-1:
                ws['I'+str(newIndex)] = invoiceObject['invoice'+str(counter)]['campaignData']['invoice' + str(counter)]["campaign" + str(campaignCounter)]["campaignImp"]
            else:
                cpcIndex = invoiceObject['invoice'+str(counter)]['campaignData']['invoice' + str(counter)]["campaign" + str(campaignCounter)]["campaignImp"].find("CPC")
                clickString = invoiceObject['invoice'+str(counter)]['campaignData']['invoice' + str(counter)]["campaign" + str(campaignCounter)]["campaignImp"]
                ws['J'+str(newIndex)] =   clickString[0:cpcIndex]
            ws['K'+str(newIndex)] =   "-------same invoice-------"
            ws['L'+str(newIndex)] =   "-------same invoice-------"
            ws['M'+str(newIndex)] =   "-------same invoice-------"
            campaignCounter = campaignCounter + 1
        ws['L'+str(newIndex)] =  invoiceObject['invoice'+str(counter)]['invoiceTax']
        ws['M'+str(newIndex)] =  invoiceObject['invoice'+str(counter)]['fileName']
        ws['K'+str(newIndex)] =  invoiceObject['invoice'+str(counter)]['invoiceAmount']

        counter = counter + 1
        #print "Index: ", index

def dbmExcelSummary(invoiceObject):
    wb.create_sheet(index=2, title="DBM Summary Invoices")
    ws = wb.get_sheet_by_name("DBM Summary Invoices")
    #Print Titles#
    ws['A1'] = "DBM Invoice Number"
    ws['B1'] = "DBM Invoice File Name"
    ws['C1'] = "Transaction ID"
    ws['D1'] = "Advertiser Name"
    ws['E1'] = "Advertiser ID"
    ws['F1'] = "Insertion Order Name"
    ws['G1'] = "Insertion Order ID"
    ws['H1'] = "Invoice Tax Amount"
    ws['I1'] = "Invoice Total Amount"
    #print "Invoice DBM Object -", invoiceObject

    #ws['A1'] = 42
    #ws['A2'] = datetime.datetime.now()
    counter = 0
    cIndex = 0
    while counter < len(invoiceObject.keys()) :
        index = counter + 2
        ws['A'+str(index)] = counter + 1
        ws['B'+str(index)] = invoiceObject['invoice'+str(counter)]['fileName']
        ws['C'+str(index)] =  invoiceObject['invoice'+str(counter)]['transactionId']
        ws['D'+str(index)] =  invoiceObject['invoice'+str(counter)]['advertiserId']
        ws['E'+str(index)] =  invoiceObject['invoice'+str(counter)]['advertiserName']
        if len(invoiceObject['invoice'+str(counter)]['campaignData'].keys()) != 0:
            ws['F'+str(index)] =  invoiceObject['invoice'+str(counter)]['campaignData']['campaign0']['campaignName']
            ws['G'+str(index)] =  invoiceObject['invoice'+str(counter)]['campaignData']['campaign0']['campaignId']
        else:
            ws['F'+str(index)] =  "--------------------"
            ws['G'+str(index)] =  "--------------------"
            ws['D'+str(index)] = "Rebilled Invoice"
            ws['E'+str(index)] = "Rebilled Invoice"
        ws['H'+str(index)] =  invoiceObject['invoice'+str(counter)]['invoiceTax']
        ws['I'+str(index)] =  invoiceObject['invoice'+str(counter)]['invoiceAmount']
        counter = counter + 1

def excelFiscal(invoiceObject):
    wb.create_sheet(index=3, title="DoubleClick Fiscal Invoices")
    ws = wb.get_sheet_by_name("DoubleClick Fiscal Invoices")
    #Print Titles#
    ws['A1'] = "Fiscal Invoice Number"
    ws['B1'] = "Fiscal Invoice File Name"
    ws['C1'] = "Transaction ID"
    ws['D1'] = "Invoice Id"
    ws['E1'] = "Invoice Subtotal Amount"
    ws['F1'] = "Invoice Total Amount"
    #ws['A1'] = 42
    #ws['A2'] = datetime.datetime.now()
    counter = 0
    while counter < len(invoiceObject.keys()) :
        index = counter + 2
        ws['A'+str(index)] = counter + 1
        ws['B'+str(index)] = invoiceObject['invoice'+str(counter)]['fileName']
        ws['C'+str(index)] =  invoiceObject['invoice'+str(counter)]['transactionId']
        ws['D'+str(index)] =  invoiceObject['invoice'+str(counter)]['facturaId']
        ws['E'+str(index)] =  invoiceObject['invoice'+str(counter)]['invoiceTax']
        ws['F'+str(index)] =  invoiceObject['invoice'+str(counter)]['invoiceTotal']
        counter = counter + 1

def dbmMatchedExcel(invoiceObject) :
    wb.create_sheet(index=4, title="DBM Matched Invoices")
    ws = wb.get_sheet_by_name("DBM Matched Invoices")
    #Print Titles#
    ws['A1'] = "Invoice Number"
    ws['B1'] = "Date"
    ws['C1'] = "Product"
    ws['D1'] = "Advertiser Name"
    ws['E1'] = "Advertiser ID"
    ws['F1'] = "Summary Transaction ID"
    ws['G1'] = "Fiscal Transaction ID"
    ws['H1'] = "Summary Invoice Subtotal Amount"
    ws['I1'] = "Summary Invoice Total Amount"
    ws['J1'] = "Fiscal Invoice Tax Amount"
    ws['K1'] = "Fiscal Invoice Total Amount"
    ws['L1'] = "Fiscal Invoice File Name"
    ws['M1'] = "Summary Invoice File Name"
    #ws['A2'] = datetime.datetime.now()
    counter = 0
    while counter < len(invoiceObject.keys()) :
        index = counter + 2
        ws['A'+str(index)] = counter + 1
        ws['B'+str(index)] = invoiceObject['invoice'+str(counter)]["summaryInvoice"]['date']
        ws['C'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['product']
        if invoiceObject['invoice'+str(counter)]["summaryInvoice"]['advertiserName'] != '':
            ws['D'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['advertiserName']
            ws['E'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['advertiserId']
        else:
            ws['D'+str(index)] = "Rebilled Invoice"
            ws['E'+str(index)] = "Rebilled Invoice"
        ws['F'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['transactionId']
        ws['G'+str(index)] =  invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['transactionId']
        ws['H'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['invoiceAmount']
        ws['I'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['invoiceTax']
        ws['J'+str(index)] = invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['invoiceTax']
        ws['K'+str(index)] =  invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['invoiceTotal']
        ws['L'+ str(index)] =  invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['fileName']
        ws['M' + str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['fileName']
        ws['N' + str(index)] =  invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['facturaId']
        counter = counter + 1

def dcmMatchedExcel(invoiceObject) :
    wb.create_sheet(index=5, title="DCM Matched Invoices")
    ws = wb.get_sheet_by_name("DCM Matched Invoices")
    #Print Titles#
    ws['A1'] = "Invoice Number"
    ws['B1'] = "Date"
    ws['C1'] = "Product"
    ws['D1'] = "Advertiser Name"
    ws['E1'] = "Advertiser ID"
    ws['F1'] = "Summary Transaction ID"
    ws['G1'] = "Fiscal Transaction ID"
    ws['H1'] = "Summary Invoice Subtotal Amount"
    ws['I1'] = "Summary Invoice Total Amount"
    ws['J1'] = "Fiscal Invoice Tax Amount"
    ws['K1'] = "Fiscal Invoice Total Amount"
    ws['L1'] = "Fiscal Invoice File Name"
    ws['M1'] = "Summary Invoice File Name"
    #ws['A2'] = datetime.datetime.now()
    counter = 0
    while counter < len(invoiceObject.keys()) :
        index = counter + 2
        ws['A'+str(index)] = counter + 1
        ws['B'+str(index)] = invoiceObject['invoice'+str(counter)]["summaryInvoice"]['date']
        ws['C'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['product']
        if invoiceObject['invoice'+str(counter)]["summaryInvoice"]['advertiserName'] != '':
            ws['D'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['advertiserName']
            ws['E'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['advertiserId']
        else:
            ws['D'+str(index)] = "Rebilled Invoice"
            ws['E'+str(index)] = "Rebilled Invoice"
        ws['F'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['transactionId']
        ws['G'+str(index)] =  invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['transactionId']
        ws['H'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['invoiceAmount']
        ws['I'+str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['invoiceTax']
        ws['J'+str(index)] = invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['invoiceTax']
        ws['K'+str(index)] =  invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['invoiceTotal']
        ws['L'+ str(index)] =  invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['fileName']
        ws['M' + str(index)] =  invoiceObject['invoice'+str(counter)]["summaryInvoice"]['fileName']
        ws['N' + str(index)] =  invoiceObject['invoice'+str(counter)]["fiscalInvoice"]['facturaId']
        counter = counter + 1

def saveExcel (excelName):
    wb.save("reports/"+excelName+".xlsx")
